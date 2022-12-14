"""

Este documento calcula el precio promedio de las otras marcas en las otras tiendas en la misma semana, 
es decir, el llamado Instrumento de Hausman. La fusión posterior de ambos archivos se 
hizo mediante la función BUSCARV() de Excel.

Requisito: Es mandatorio que el excel esté previamente ordenado, primero por marca, de menor a mayor
y luego por semana, también de menor a mayor. Futuras versiones podrían evitar este requisito.

input: DATA_UDESA.xlsx
output: hausman_instrument.xlsx


"""
import os
import pandas as pd
from openpyxl import load_workbook 

#Seteamos directorio
os.chdir('C:\\Users\\Franco\\Desktop\\UDESA\\Metodos econométricos e IO aplicada\\Examen Final')

#Creamos libro

libro = load_workbook('input/DATA_UDESA.xlsx') 

hoja = libro['DATA de Medicamentos'] #Indicamos la hoja a escribir

row = 1
col = 10 

hoja.cell(row, col).value = 'Hausman' #titulo de la columna
row += 1


df0 = pd.read_excel("input/DATA_UDESA.xlsx")

df00 = df0.sort_values(by = ['semana', 'marca', 'tienda'])

lista_df = []

for i in range(1, 49):    #Creamos un loop para armar una lista de data frames,
                          #donde cada data frame contiene todos los datos correspondientes
                          #a una semana (en total 48 data frames en la lista).

    w_i = df00[(df00['semana'] == i)]
    w_i.reset_index(inplace = True)
    lista_df.append(w_i)




for i in range(len(lista_df)):      #Para cada data frame (semana) en la lista;
   
    for x in range(len(lista_df[i])):   #Para cada observacion (marca tienda) en la lista;
        suma = 0
        cuanto_divido = 0

        for ele in range(len(lista_df[i])):  #Chequeamos las observaciones de esa semana y...

            if lista_df[i].iloc[x]['marca'] != lista_df[i].iloc[ele]['marca']: #...si es otra marca...

                if lista_df[i].iloc[x]['tienda'] != lista_df[i].iloc[ele]['tienda']: #...si es otra tienda...

                    suma += lista_df[i].iloc[ele]['precio'] #...incluimos su precio
                    cuanto_divido += 1                      #y aumentamos el divisor del promedio.

        hoja.cell(row, col).value = suma/cuanto_divido #Finalmente, escribimos en la hoja el precio promedio
                                                       #de las otras marca tienda en esa semana (inst. de Hausman)
        row += 1
    
    
libro.save('input/DATA_UDESA.xlsx')

