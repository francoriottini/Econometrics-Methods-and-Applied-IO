"""

Calculador de precios promedios por marca y semana.

Requiere como requisito previo que los datos esten ordenados por semana y marca.

input: DATA_UDESA.xlsx
output: DATA_UDESA.xlsx


"""

import os
import pandas as pd                                               
from openpyxl import load_workbook


#Seteamos directorio
os.chdir('C:\\Users\\Franco\\Desktop\\UDESA\\Metodos econométricos e IO aplicada\\Examen Final')

libro = load_workbook('input/DATA_UDESA.xlsx') #Importamos el Excel con los datos
                                                                    #(ORDENADOS PREVIAMENTE POR SEMANA, MARCA Y TIENDA)
hoja = libro['DATA de Medicamentos'] #Indicamos la hoja a escribir

row = 1
col = 12 

hoja.cell(row, col).value = 'precio_jt' #Colocamos el título a la columna que tendrá el instrumento de Hausman.
row += 1


df1 = pd.read_excel("input/DATA_UDESA.xlsx")
df1 = df1.sort_values(by = ['semana','marca'])

df2 = df1.groupby(by = ['semana','marca']).agg({'precio':'mean'})
df2.reset_index(inplace = True)

cuenta=0
lista_totales=[]
for i in range(len(df2)):

        while (hoja.cell(row, 3).value == df2.iloc[i]['marca']) and (hoja.cell(row, 2).value == df2.iloc[i]['semana']): #Como son shares semanales, se repiten
                                                                 #en todas las tiendas donde se venda
                                                                 #esa marca en esa semana.
            hoja.cell(row, col).value = (df2.iloc[i]['precio'])
            row += 1



libro.save('input/DATA_UDESA.xlsx') 

























