"""

Este documento calcula los share de venta de cada marca en cada semana, es decir,
los s_{jt}. La fusión posterior de ambos archivos se hizo mediante la función
BUSCARV() de Excel.

Requisito: Es mandatorio que el excel esté previamente ordenado, primero por marca, de menor a mayor
y luego por semana, también de menor a mayor. Futuras versiones podrían evitar este requisito.

input: DATA_UDESA.xlsx
output: Shares.xlsx


"""
import os
import pandas as pd
from openpyxl import load_workbook 

#Seteamos directorio
os.chdir('C:\\Users\\Franco\\Desktop\\UDESA\\Metodos econométricos e IO aplicada\\Examen Final')

#Input
libro = load_workbook('input\DATA_UDESA.xlsx')

hoja = libro['DATA de Medicamentos'] #Indicamos la hoja a escribir

row = 1
col = 9 

hoja.cell(row, col).value = 'share_jt' #Nombre de la columna
row += 1

df = pd.read_excel("input\DATA_UDESA.xlsx") #Importamos los datos y los colocamos en un data frame.

df2 = df.groupby(by = ['semana', 'marca']).agg({'ventas':'sum'}) #Esta lÃ­nea calcula las ventas de cada 
                                                                 #marca por semana (sin distincion de tienda)
df2.reset_index(inplace = True)                                  #en un nuevo data frame

df3 = df2.groupby(by = ['semana']).agg({'ventas':'sum'}) #Esta li­nea calcula las ventas TOTALES por semana
 
df3.reset_index(inplace = True)
cuenta=0
lista_totales=[]
#Este for loop calcula el share semanal de cada marca.
for i in range(48): 	                                   
	for ele in range(cuenta, cuenta + 11):
		while hoja.cell(row, 3).value == df2.iloc[ele]['marca']: #Como son shares semanales, se repiten
                                                                 #en todas las tiendas donde se venda
                                                                 #esa marca en esa semana.
			hoja.cell(row, col).value = (df2.iloc[ele]['ventas']/df3.iloc[i]['ventas'])*0.62
			row += 1
	cuenta = cuenta + 11 #Actualizamos el contador para pasar a la siguiente semana.


libro.save('input\DATA_UDESA.xlsx') #Guardamos los cambios.





