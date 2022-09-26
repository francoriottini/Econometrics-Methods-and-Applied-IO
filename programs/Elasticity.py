"""
Calculador de las elasticidades promedio anual para los modelos logit 1, 2 y 3.

Input: Demand_Estimation.xlsx con las elasticidades individuales ya calculadas (SALIDA de STATA), en 
                       columnas con el título "elasticity1", "elasticity2" y "elasticity3"

Requiere: Pandas y Openpyxl instalados.

Lucas Castellini y Franco Riottini
"""
import os
import pandas as pd                          
               
from openpyxl import load_workbook           

#Seteamos directorio
os.chdir('C:\\Users\\Franco\\Desktop\\UDESA\\Metodos econométricos e IO aplicada\\Examen Final')

#Importamos el Excel con los datos y las elasticidades previamente calculadas.
libro = load_workbook('output/Demand_Estimation.xlsx') 


#Indicamos la hoja a escribir
hoja = libro['Sheet1'] 

#Indicamos la columnas
row = 1
col = 21 

#Colocamos los títulos de las columnas donde se imprimiran las elasticidades medias.
hoja.cell(row, col).value = ' M1 mean elasticity_j'
col += 1
hoja.cell(row, col).value = ' M2 mean elasticity_j'
col += 1
hoja.cell(row, col).value = ' M3 mean elasticity_j'

col=21
row +=1

#Importamos nuevamente los datos para crear un data frame en Pandas.
df = pd.read_excel("output/Demand_Estimation.xlsx") 

#Data frame para elasticidad promedio 1.
df2 = df.groupby(by = ['marca']).agg({'elasticity1':'mean'})                                                                  
df2.reset_index(inplace = True)                                  

#Data frame para elasticidad promedio 2.
df3 = df.groupby(by = ['marca']).agg({'elasticity2':'mean'})                                                                  
df3.reset_index(inplace = True)

#Data frame para elasticidad promedio 3.
df4 = df.groupby(by = ['marca']).agg({'elasticity3':'mean'})                                                                   
df4.reset_index(inplace = True)


dicc = {} 
dicc2 = {}
dicc3 = {}

for ele in range(len(df2)): #For loop que rellena diccionarios, 
                            #mapeando marcas con elasticidades medias.

	dicc[df2.iloc[ele]['marca']] = round(df2.iloc[ele]['elasticity1'], 4)
	dicc2[df2.iloc[ele]['marca']] = round(df3.iloc[ele]['elasticity2'], 4)
	dicc3[df2.iloc[ele]['marca']] = round(df4.iloc[ele]['elasticity3'], 4)


	  

s = pd.DataFrame(dicc.values())
s['(1)'] = dicc.values()
s['(2)'] = dicc2.values()
s['(3)'] = dicc3.values()

s.index = ['Marca 1 (25)','Marca 1 (50)','Marca 1 (100)','Marca 2 (25)','Marca 2 (50)',

            'Marca 2 (100)','Marca 3 (25)','Marca 3 (50)','Marca 3 (100)','Marca 4 (50)',

            'Marca 4 (100)']

s = s.drop([0], axis = 1)


           
s.to_latex('output/Elasticidades.tex', index=True, caption = 'Elasticidades medias anuales')


