"""
Calculador de precios promedios por marca j y semana t (precio_jt)

Input: DATA_UDESA.xlsx ORDENADO POR SEMANA Y MARCA ( semana1:marca1, semana1:marca2, ..., semana48:marca11)

Demora: menos de 30 segundos

Requiere: Pandas y Openpyxl instalados.

Lucas Castellini y Franco Riottini
"""
import os
import pandas as pd                                               
from openpyxl import load_workbook


#Seteamos directorio
os.chdir('C:\\Users\\Franco\\Desktop\\UDESA\\Metodos econométricos e IO aplicada\\Examen Final')

#Importamos el Excel con los datos (ORDENADO PREVIAMENTE POR SEMANA:MARCA)
libro = load_workbook('input/DATA_UDESA.xlsx') 
                                                                   

#Indicamos la hoja a escribir
hoja = libro['DATA de Medicamentos'] 

#Indicamos fila y columna
row = 1  
col = 11 

#Colocamos el título a la columna.
hoja.cell(row, col).value = 'precio_jt' 
row += 1

#Importamos nuevamente los datos para crear un data frame de Pandas.
df1 = pd.read_excel("input/DATA_UDESA.xlsx")
df1 = df1.sort_values(by = ['semana','marca'])

df2 = df1.groupby(by = ['marca']).agg({'precio':'mean'})
df2.reset_index(inplace = True)

df3 = df1.groupby(by = ['marca']).agg({'costo':'mean'})
df3.reset_index(inplace = True)

lista = []

for ele in range(len(df3)):

    lista.append(str(round(df3.iloc[ele]['costo']/df2.iloc[ele]['precio'],2))+' %')


s = pd.Series(lista)


s.index = ['Marca 1','Marca 2','Marca 3','Marca 4','Marca 5','Marca 6','Marca 7','Marca 8','Marca 9','Marca 10','Marca 11']

s.to_latex('output/markup.tex', header = ['Mark-up'], caption = ('Mark-up promedio anual'))




